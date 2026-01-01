import os
import smtplib
from email.message import EmailMessage
from email.utils import formataddr
from dotenv import load_dotenv
from openpyxl import load_workbook

# Load environment variables
load_dotenv()

SENDER_EMAIL = os.getenv("EMAIL")
PASSWORD = os.getenv("PASSWORD")

if not SENDER_EMAIL or not PASSWORD:
    raise ValueError("EMAIL or PASSWORD not found in .env")

SMTP_SERVER = "smtp.gmail.com"
PORT = 587

# Load Excel data
wb = load_workbook("customers.xlsx")
sheet = wb.active

def send_email(name, receiver_email, loan_id, due_date, amount):
    msg = EmailMessage()
    msg["Subject"] = "Loan Payment Reminder"
    msg["From"] = formataddr(("Loan Department", SENDER_EMAIL))
    msg["To"] = receiver_email

    msg.set_content(
        f"""Dear {name},

This is a reminder that your remaining loan amount of {amount} INR
for Loan ID {loan_id} is due on {due_date}.

Please make the payment at the earliest.

Regards,
Loan Department
"""
    )

    with smtplib.SMTP(SMTP_SERVER, PORT) as server:
        server.starttls()
        server.login(SENDER_EMAIL, PASSWORD)
        server.send_message(msg)

# Iterate through Excel rows
for row in sheet.iter_rows(min_row=2, values_only=True):
    name, email, loan_id, due_date, amount = row
    send_email(name, email, loan_id, due_date, amount)
    print(f"Email sent to {name} ({email})")

print("All reminder emails sent successfully.")